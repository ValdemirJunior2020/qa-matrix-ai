from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

ALLOWED_EXTENSIONS = {".xlsx"}

def validate_xlsx(path: Path, max_mb: int = 20) -> dict:
    if path.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise ValueError("Only .xlsx files are allowed")
    if path.stat().st_size > max_mb * 1024 * 1024:
        raise ValueError(f"Matrix exceeds {max_mb} MB limit")
    with path.open("rb") as f:
        sig = f.read(4)
    if sig != b"PK\x03\x04":
        raise ValueError("File is not a valid XLSX container")
    wb = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    return {"sheet_names": wb.sheetnames, "sheet_count": len(wb.sheetnames)}
