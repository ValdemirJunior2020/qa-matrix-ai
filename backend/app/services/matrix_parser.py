from __future__ import annotations
import hashlib, json, re, uuid
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HEADER_WORDS = {"instructions","instruction","slack","refund queue","create a ticket","supervisor","vipres","score","critical"}

def _txt(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def _range_for_row(ws, row: int, min_col: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{row}:{get_column_letter(max_col)}{row}"

def _looks_like_section(values: list[str | None]) -> bool:
    clean = [str(x).strip().lower() for x in values if x]
    if len(clean) < 2: return False
    hits = sum(1 for x in clean[1:] if x in HEADER_WORDS)
    return hits >= 1

def parse_matrix(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, keep_links=False)
    records = []
    sheet_meta = []
    for ws in wb.worksheets:
        dims = ws.calculate_dimension()
        merged = [str(r) for r in ws.merged_cells.ranges]
        hidden = ws.sheet_state != "visible"
        sheet_meta.append({"name": ws.title, "dimension": dims, "hidden": hidden, "merged_ranges": merged})
        if hidden:
            continue
        current_category = None
        current_headers = {}
        for r in range(1, ws.max_row + 1):
            values = [_txt(ws.cell(r,c).value) for c in range(1, ws.max_column + 1)]
            if not any(values):
                continue
            if _looks_like_section(values):
                current_category = next((v for v in values if v), None)
                current_headers = {c: (_txt(ws.cell(r,c).value) or f"Column {get_column_letter(c)}") for c in range(1, ws.max_column + 1)}
                continue
            first_data_col = next((c for c in range(1, ws.max_column + 1) if _txt(ws.cell(r,c).value)), None)
            if not first_data_col:
                continue
            # Workbook-specific rows usually start in column B; generic fallback supports any first populated column.
            rule = _txt(ws.cell(r, 2).value) if ws.max_column >= 2 else None
            if not rule:
                rule = _txt(ws.cell(r, first_data_col).value)
            # Skip notes/title rows that are not under a section but keep them as policy notes.
            category = current_category or "General / Notes"
            fields = {}
            for c in range(1, ws.max_column + 1):
                val = _txt(ws.cell(r,c).value)
                if val is None: continue
                key = current_headers.get(c) or f"Column {get_column_letter(c)}"
                fields[key] = val
            instructions = None
            for k,v in fields.items():
                if "instruction" in k.lower():
                    instructions = v; break
            score = None
            critical = False
            critical_condition = None
            for k,v in fields.items():
                kl = k.lower(); vl = v.lower()
                if "score" in kl:
                    m = re.search(r"-?\d+(?:\.\d+)?", v.replace(",",""))
                    if m: score = float(m.group())
                if "critical" in kl:
                    if vl in {"yes","y","true","critical"} or "critical" in vl:
                        critical = True; critical_condition = v
            comments = {get_column_letter(c): ws.cell(r,c).comment.text for c in range(1, ws.max_column + 1) if ws.cell(r,c).comment}
            formula_cells = {get_column_letter(c): str(ws.cell(r,c).value) for c in range(1, ws.max_column + 1) if isinstance(ws.cell(r,c).value, str) and ws.cell(r,c).value.startswith("=")}
            style_hints = {get_column_letter(c): {"number_format": ws.cell(r,c).number_format, "bold": bool(ws.cell(r,c).font and ws.cell(r,c).font.bold)} for c in range(1, ws.max_column + 1) if _txt(ws.cell(r,c).value)}
            raw_join = " | ".join(f"{k}: {v}" for k,v in fields.items())
            rec = {
                "id": str(uuid.uuid5(uuid.NAMESPACE_URL, f"{path.name}|{ws.title}|{r}|{raw_join}")),
                "workbook": path.name,
                "sheet": ws.title,
                "category": category,
                "subcategory": rule,
                "rule": rule,
                "instructions": instructions,
                "fields": fields,
                "comments": comments,
                "formula_cells": formula_cells,
                "style_hints": style_hints,
                "source_row_start": r,
                "source_row_end": r,
                "cell_range": _range_for_row(ws, r, first_data_col, ws.max_column),
                "score": score,
                "critical": critical,
                "critical_condition": critical_condition,
                "raw_text": raw_join,
            }
            records.append(rec)
    sha = hashlib.sha256(path.read_bytes()).hexdigest()
    return {"filename": path.name, "sha256": sha, "sheet_count": len(wb.sheetnames), "sheets": sheet_meta, "records": records, "rule_count": len(records), "named_ranges": [str(x) for x in wb.defined_names.values()]}
